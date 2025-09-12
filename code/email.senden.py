import smtplib
import ssl
import getpass
from email.message import EmailMessage
from time import sleep
from openpyxl import load_workbook

# --- KONFIGURATION: Bitte passe diesen Teil an ---

SMTP_SERVER = "mail.bbw.ch"
SMTP_PORT = 465
EMAIL_SENDER = "pietro.rossi@bbw.ch"
XLSX_FILENAME = "data.xlsx"

EMAIL_SUBJECT = "Ihr persönlicher Code ist da!"
EMAIL_BODY_TEMPLATE = """
Hallo,

vielen Dank für Ihre Teilnahme!

Ihr persönlicher Code lautet: {code}

Bitte bewahren Sie diesen Code gut auf.

Viele Grüße,
Ihr Team
"""

# --- Ende der Konfiguration ---


def send_personalized_emails():
    try:
        email_password = getpass.getpass(f"Bitte Passwort für {EMAIL_SENDER} eingeben: ")
    except Exception as e:
        print(f"Konnte Passwort nicht lesen: {e}")
        return

    context = ssl.create_default_context()

    try:
        print(f"Verbinde mit Server {SMTP_SERVER}...")
        with smtplib.SMTP_SSL(SMTP_SERVER, SMTP_PORT, context=context) as server:
            print("Erfolgreich verbunden. Logge ein...")
            server.login(EMAIL_SENDER, email_password)
            print("Login erfolgreich. Beginne mit dem E-Mail-Versand.")

            workbook = load_workbook(filename=XLSX_FILENAME)
            sheet = workbook.active

            for row in sheet.iter_rows(min_row=2, values_only=True):
                try:
                    recipient_email, code = row

                    msg = EmailMessage()
                    msg['Subject'] = EMAIL_SUBJECT
                    msg['From'] = EMAIL_SENDER
                    msg['To'] = recipient_email

                    body = EMAIL_BODY_TEMPLATE.format(code=code, email=recipient_email)
                    msg.set_content(body)

                    server.send_message(msg)
                    print(f"✅ E-Mail an {recipient_email} erfolgreich gesendet.")

                    sleep(1)

                except Exception as e:
                    print(f"❌ Fehler beim Senden an {row[0]}: {e}")

    except FileNotFoundError:
        print(f"FEHLER: Die Datei '{XLSX_FILENAME}' wurde nicht gefunden. Stelle sicher, dass sie im selben Ordner liegt.")
    except smtplib.SMTPAuthenticationError:
        print("FEHLER: Login fehlgeschlagen. Überprüfe E-Mail-Adresse und Passwort (oder App-Passwort).")
    except Exception as e:
        print(f"Ein unerwarteter Fehler ist aufgetreten: {e}")

    print("\nSkript beendet.")


if __name__ == "__main__":
    send_personalized_emails()
